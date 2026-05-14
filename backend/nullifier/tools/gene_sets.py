"""Gene-set expansion for Nullifier v6.

Sources:
  * **SynGO 1.3** — local Excel dump at ``syngo1.3_complete_data/`` (already in the repo).
    Three sheets: ``genes.xlsx`` (master SynGO gene list with HGNC/Ensembl/Entrez ids),
    ``annotations.xlsx`` (gene→GO annotations with evidence), ``ontologies.xlsx`` (the
    304 GO terms SynGO uses, with each term's full member list in the ``hgnc_symbol``
    column).  Release stamped ``"1.3"``.

    Reference: Koopmans et al., *Neuron* 2019, 103(2):217–234.

  * **BBB sets** — hardcoded marker panels for the major BBB cell types. Sourced from
    Vanlandewijck et al., *Nature* 2018 ("A molecular atlas of cell types and zonation
    in the brain vasculature") and the Allen Brain Atlas brain-vasculature cell-type
    markers. These are short, curated lists — not the full transcriptome — so the
    Methodologist has a tractable comparison set.

  * **Control set** — a small, hardcoded panel of widely-expressed brain housekeeping
    genes, matched on broad criteria (constitutively expressed, no overt synaptic or
    vascular specialisation). The matching is qualitative; the matching parameters are
    surfaced in the expansion record so the UI can show them and the Skeptic can
    critique them.

This module never calls Ensembl. The relevance classifier (which canonical sets to
include for *this* hypothesis) routes through ``llm_call_json`` with the
``gene_set_classifier`` routing key, which the user can point to local Gemma.
"""
from __future__ import annotations

import pickle
import time
from pathlib import Path
from typing import Iterable

import openpyxl

from ..provenance import make_provenance
from ..config.loader import load_config

SYNGO_RELEASE = "1.3"
SYNGO_DIR = Path(__file__).resolve().parents[3] / "syngo1.3_complete_data"
SYNGO_CACHE = Path.home() / ".nullifier" / "gene_sets_cache.pkl"

# ── BBB cell-type marker panels (curated; cite in the panel) ────────────────
# Endothelial: classical brain-EC markers (tight junctions, transporters, selectivity).
# Astrocyte (support): astrocyte end-foot / BBB-supporting markers.
# Pericyte: mural/pericyte markers.
BBB_VERSION = "v1 (Vanlandewijck 2018 + Allen Brain Atlas curated)"

BBB_SETS: dict[str, list[str]] = {
    "endothelial":     ["CLDN5", "OCLN", "TJP1", "TJP2", "CDH5", "PECAM1",
                        "SLC2A1", "ABCB1", "ABCG2", "MFSD2A", "LSR", "JAM3",
                        "PLVAP", "VWF", "FLT1", "KDR", "TEK", "ANGPT2",
                        "INSR", "LRP1", "TFRC", "BSG", "PGCP", "ICAM2"],
    "astrocyte_support": ["GFAP", "AQP4", "S100B", "ALDH1L1", "SLC1A2", "SLC1A3",
                          "GJA1", "AGT", "VIM", "FGFR3", "MLC1", "ATP1B2",
                          "GLUL", "APOE", "CLU"],
    "pericyte":        ["PDGFRB", "MCAM", "RGS5", "CSPG4", "DES", "ACTA2",
                        "MYH11", "NOTCH3", "ANPEP", "KCNJ8", "ABCC9", "P2RX1"],
}

# ── Control set: brain-expressed housekeeping / reference panel ─────────────
# Picked for broad, stable expression in CNS tissues; no overt synaptic or vascular
# specialisation. Matching is qualitative — the parameters below describe the lens.
CONTROL_MATCHING = {
    "tissue_expression":  "broadly expressed across CNS cell types",
    "gene_length":        "not size-selected (range)",
    "gc_content":         "not GC-selected (range)",
    "constraint_pli":     "not constraint-selected (mixed)",
    "tissue_specificity": "low specificity (housekeeping-like)",
    "note":               "Hardcoded curated panel; not a statistically matched control. "
                          "Treat as a qualitative reference, not a null distribution.",
}

CONTROL_SETS: dict[str, list[str]] = {
    "housekeeping_brain": [
        "GAPDH", "ACTB", "B2M", "PGK1", "HPRT1", "TBP", "PPIA", "RPL13A",
        "YWHAZ", "UBC", "HMBS", "GUSB", "SDHA", "TFRC", "EEF1A1", "PUM1",
    ],
}


# ── SynGO parser ────────────────────────────────────────────────────────────
def _parse_syngo(data_dir: Path) -> dict:
    """Parse the three SynGO sheets. Returns ``{release, genes, annotations, ontologies}``
    where ``ontologies`` is keyed by GO id and includes a ``members`` list (HGNC symbols)."""
    wb_g = openpyxl.load_workbook(data_dir / "genes.xlsx", read_only=True, data_only=True)
    ws = wb_g.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}
    genes = []
    for row in rows:
        genes.append({
            "hgnc_id":     row[col["hgnc_id"]],
            "hgnc_symbol": row[col["hgnc_symbol"]],
            "ensembl_id":  row[col["ensembl_id"]],
            "entrez_id":   row[col["entrez_id"]],
        })
    wb_g.close()

    wb_a = openpyxl.load_workbook(data_dir / "annotations.xlsx", read_only=True, data_only=True)
    ws = wb_a.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}
    annotations: list[dict] = []
    for row in rows:
        annotations.append({
            "hgnc_symbol": row[col["hgnc_symbol"]],
            "go_id":       row[col["go_id"]],
            "go_name":     row[col["go_name"]],
            "go_domain":   row[col["go_domain"]],
        })
    wb_a.close()

    wb_o = openpyxl.load_workbook(data_dir / "ontologies.xlsx", read_only=True, data_only=True)
    ws = wb_o.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}
    ontologies: dict[str, dict] = {}
    for row in rows:
        go_id = row[col["id"]]
        if not go_id:
            continue
        members_raw = row[col["hgnc_symbol"]] or ""
        members = sorted({m.strip() for m in str(members_raw).split(",") if m.strip()})
        ontologies[go_id] = {
            "id":        go_id,
            "domain":    row[col["domain"]],
            "name":      row[col["name"]],
            "shortname": row[col["shortname"]],
            "parent_id": row[col["parent_id"]],
            "members":   members,
        }
    wb_o.close()
    return {"release": SYNGO_RELEASE, "genes": genes, "annotations": annotations,
            "ontologies": ontologies}


def load_syngo(data_dir: Path | None = None, use_cache: bool = True,
               cache_ttl_days: int = 7) -> dict:
    """Load (and cache) the parsed SynGO dump. Re-parses if the cache is older than
    the cache TTL or any of the source ``.xlsx`` files has changed since the cache
    was written."""
    data_dir = Path(data_dir or SYNGO_DIR)
    if not data_dir.exists():
        raise FileNotFoundError(f"SynGO data directory not found: {data_dir}")
    sources = [data_dir / "genes.xlsx", data_dir / "annotations.xlsx",
               data_dir / "ontologies.xlsx"]
    for p in sources:
        if not p.exists():
            raise FileNotFoundError(f"SynGO file missing: {p}")
    SYNGO_CACHE.parent.mkdir(parents=True, exist_ok=True)
    if use_cache and SYNGO_CACHE.exists():
        cache_age = time.time() - SYNGO_CACHE.stat().st_mtime
        max_src_mtime = max(p.stat().st_mtime for p in sources)
        if cache_age < cache_ttl_days * 86400 and SYNGO_CACHE.stat().st_mtime >= max_src_mtime:
            try:
                with open(SYNGO_CACHE, "rb") as f:
                    return pickle.load(f)
            except Exception:
                pass  # fall through to a fresh parse
    parsed = _parse_syngo(data_dir)
    try:
        with open(SYNGO_CACHE, "wb") as f:
            pickle.dump(parsed, f)
    except Exception:
        pass
    return parsed


# ── Canonical set assembly from SynGO ───────────────────────────────────────
# Top-level SynGO terms we surface as candidate sets. Other BP terms come along as
# `synaptic.process.<shortname>` derived from ontologies.xlsx.
SYNGO_TOPLEVEL = {
    "synaptic.all":          "GO:0045202",   # synapse (CC root)
    "synaptic.presynaptic":  "GO:0098793",   # presynapse
    "synaptic.postsynaptic": "GO:0098794",   # postsynapse
}


def _syngo_canonical_sets(syngo: dict) -> dict[str, list[str]]:
    sets: dict[str, list[str]] = {}
    onto = syngo["ontologies"]
    for label, go_id in SYNGO_TOPLEVEL.items():
        if go_id in onto:
            sets[label] = list(onto[go_id]["members"])
    # BP-domain processes — one set per term, prefixed `synaptic.process.<shortname>`
    for go_id, term in onto.items():
        if term["domain"] == "BP" and term["members"]:
            short = (term["shortname"] or term["name"] or go_id).strip().lower().replace(" ", "_")
            sets[f"synaptic.process.{short}"] = list(term["members"])
    return sets


def _all_canonical_sets(syngo: dict) -> dict[str, dict]:
    """name -> {genes, source, label}; merges SynGO + BBB + control sets."""
    out: dict[str, dict] = {}
    for name, genes in _syngo_canonical_sets(syngo).items():
        out[name] = {"genes": genes, "source": f"SynGO {SYNGO_RELEASE}", "label": name}
    for name, genes in BBB_SETS.items():
        out[f"bbb.{name}"] = {"genes": list(genes), "source": f"BBB curated {BBB_VERSION}",
                              "label": f"bbb.{name}"}
    for name, genes in CONTROL_SETS.items():
        out[f"control.{name}"] = {"genes": list(genes), "source": "Hardcoded control panel",
                                  "label": f"control.{name}"}
    return out


# ── Relevance classifier (Gemma; routes per config) ─────────────────────────
_RELEVANCE_SYSTEM = """You are filtering canonical gene sets for relevance to a scientific
hypothesis. For each candidate set you will see: the set name, a one-line description, the
size (gene count), and a small sample of member genes. Return a relevance score 0-3 where:
  0 = unrelated
  1 = peripheral (don't include unless nothing better matches)
  2 = clearly relevant
  3 = central to the hypothesis
Respond with ONLY valid JSON: {"score": 0-3, "why": "one short clause"}"""


def _gemma_relevance(hypothesis: str, candidate: dict) -> tuple[int, str]:
    """Returns ``(score, why)``. Falls back to a heuristic if the LLM call fails."""
    from .llm_client import llm_call_json  # local import: avoid cycle at module load
    sample = ", ".join(candidate["genes"][:8])
    user = (f'HYPOTHESIS: {hypothesis}\n\n'
            f'SET: {candidate["label"]}\n'
            f'SOURCE: {candidate["source"]}\n'
            f'SIZE: {len(candidate["genes"])} genes\n'
            f'SAMPLE MEMBERS: {sample}\n')
    try:
        out = llm_call_json("gene_set_classifier", _RELEVANCE_SYSTEM, user, max_tokens=120)
        score = int(out.get("score", 0))
        return max(0, min(3, score)), str(out.get("why", ""))[:200]
    except Exception as e:
        # Heuristic fallback: keep BBB/synaptic top-level + housekeeping; drop sub-processes.
        label = candidate["label"]
        if label in {"synaptic.presynaptic", "synaptic.postsynaptic", "synaptic.all"}:
            return 2, "fallback: synaptic top-level"
        if label.startswith("bbb."):
            return 2, "fallback: BBB cell type"
        if label.startswith("control."):
            return 1, "fallback: control"
        return 0, f"fallback (llm error: {type(e).__name__})"


# ── expand() — public entry point ───────────────────────────────────────────
def expand(starter_entities: Iterable[str], hypothesis: str, domain: str,
           syngo_dir: Path | None = None, min_score: int = 2,
           on_event=None) -> dict:
    """Build the v6 gene-set expansion for any domain with starter entities."""
    starter = [s.strip() for s in (starter_entities or []) if isinstance(s, str) and s.strip()]

    cfg = load_config()
    ttl = int(cfg.get("gene_sets", {}).get("cache_ttl_days", 7))
    syngo = load_syngo(syngo_dir, cache_ttl_days=ttl)
    candidates = _all_canonical_sets(syngo)

    scored: list[dict] = []
    for name, info in candidates.items():
        score, why = _gemma_relevance(hypothesis, info)
        scored.append({"set": name, "score": score, "why": why, "source": info["source"],
                       "size": len(info["genes"])})
        if on_event is not None:
            on_event(name, score)

    expanded: dict[str, list[str]] = {}
    controls: dict[str, list[str]] = {}
    starter_set = {s.upper() for s in starter}
    for s in scored:
        if s["score"] < min_score:
            continue
        name = s["set"]
        genes = candidates[name]["genes"]
        # Union with starter (no de-dup of starter — they appear in `starter` field).
        unique = [g for g in genes if g.upper() not in starter_set]
        if name.startswith("control."):
            controls[name] = unique
        else:
            expanded[name] = unique

    # Always attach a default control set if none survived scoring
    if not controls:
        for name, genes in CONTROL_SETS.items():
            controls[f"control.{name}"] = list(genes)

    total_expanded = sum(len(v) for v in expanded.values())
    total_controls = sum(len(v) for v in controls.values())

    prov = make_provenance(
        source="gene_sets.expand",
        triggered_by=["formalizer.starter_entities", "formalizer.core_hypothesis"],
        evidence_refs=[f"SynGO {SYNGO_RELEASE}", BBB_VERSION] + [s["set"] for s in scored if s["score"] >= min_score],
        method=(f"Score all canonical sets (SynGO {SYNGO_RELEASE} + BBB + controls) for "
                f"relevance to the hypothesis via the gene_set_classifier routing; keep "
                f"score >= {min_score}; union with the starter list; attach matched controls."),
        inputs={"starter": starter, "hypothesis": hypothesis, "min_score": min_score,
                "candidate_count": len(candidates)},
    )

    return {
        "skipped": False,
        "starter": starter,
        "starter_count": len(starter),
        "expanded": expanded,
        "controls": controls,
        "candidate_scores": scored,
        "min_score": min_score,
        "source": f"SynGO {SYNGO_RELEASE} + {BBB_VERSION} + hardcoded controls",
        "syngo_release": SYNGO_RELEASE,
        "bbb_version": BBB_VERSION,
        "matching_params": CONTROL_MATCHING,
        "total_expanded": total_expanded,
        "total_controls": total_controls,
        "provenance": prov,
    }


def all_genes(expansion: dict) -> list[str]:
    """Flat de-duplicated gene list across starter + expanded + controls."""
    out: list[str] = []
    seen: set[str] = set()
    for g in expansion.get("starter", []):
        if g.upper() not in seen:
            out.append(g); seen.add(g.upper())
    for genes in expansion.get("expanded", {}).values():
        for g in genes:
            if g.upper() not in seen:
                out.append(g); seen.add(g.upper())
    for genes in expansion.get("controls", {}).values():
        for g in genes:
            if g.upper() not in seen:
                out.append(g); seen.add(g.upper())
    return out


if __name__ == "__main__":
    # Smoke: parse SynGO, count sets, no LLM call.
    s = load_syngo()
    onto = s["ontologies"]
    print(f"SynGO {s['release']}: {len(s['genes'])} genes, {len(s['annotations'])} annotations, {len(onto)} GO terms")
    for label, go_id in SYNGO_TOPLEVEL.items():
        if go_id in onto:
            print(f"  {label} ({go_id}): {len(onto[go_id]['members'])} members")
    print(f"BBB sets: {[ (k, len(v)) for k,v in BBB_SETS.items() ]}")
